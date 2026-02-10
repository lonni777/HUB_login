"""
Утиліта для валідації Excel файлів мапінгу.
Містить методи для перевірки структури та даних в Excel файлах.
"""
from pathlib import Path
from typing import List, Dict, Optional, Tuple
import openpyxl
from openpyxl import load_workbook
import requests
import xml.etree.ElementTree as ET


class ExcelValidator:
    """Клас для валідації Excel файлів мапінгу"""
    
    def __init__(self, file_path: str):
        """
        Ініціалізація валідатора Excel файлу
        
        Args:
            file_path: Шлях до Excel файлу
        """
        self.file_path = Path(file_path)
        if not self.file_path.exists():
            raise FileNotFoundError(f"Excel файл не знайдено: {file_path}")
        
        self.workbook = None
        self._load_workbook()
    
    def _load_workbook(self):
        """Завантажити Excel файл"""
        try:
            self.workbook = load_workbook(self.file_path, data_only=True)
        except Exception as e:
            raise Exception(f"Помилка при завантаженні Excel файлу: {e}")
    
    def get_sheet_names(self) -> List[str]:
        """
        Отримати список назв всіх вкладок (листів) в Excel файлі
        
        Returns:
            Список назв вкладок
        """
        if not self.workbook:
            return []
        return self.workbook.sheetnames
    
    def sheet_exists(self, sheet_name: str) -> bool:
        """
        Перевірити чи існує вкладка з заданою назвою
        
        Args:
            sheet_name: Назва вкладки для перевірки
        
        Returns:
            True якщо вкладка існує, False якщо ні
        """
        if not self.workbook:
            return False
        return sheet_name in self.workbook.sheetnames
    
    def verify_sheets_exist(self, expected_sheets: List[str]) -> Tuple[bool, List[str]]:
        """
        Перевірити наявність очікуваних вкладок
        
        Args:
            expected_sheets: Список назв вкладок, які повинні бути в файлі
        
        Returns:
            Tuple (всі_вкладки_знайдені, список_відсутніх_вкладок)
        """
        if not self.workbook:
            return False, expected_sheets
        
        existing_sheets = set(self.workbook.sheetnames)
        expected_sheets_set = set(expected_sheets)
        
        missing_sheets = list(expected_sheets_set - existing_sheets)
        all_found = len(missing_sheets) == 0
        
        return all_found, missing_sheets
    
    def read_sheet_data(self, sheet_name: str, header_row: int = 1) -> List[Dict[str, any]]:
        """
        Прочитати дані з вкладки Excel файлу
        
        Args:
            sheet_name: Назва вкладки
            header_row: Номер рядка з заголовками (за замовчуванням 1)
        
        Returns:
            Список словників, де кожен словник - це рядок з даними
            Ключі словника - назви колонок з заголовків
        """
        if not self.sheet_exists(sheet_name):
            raise ValueError(f"Вкладка '{sheet_name}' не знайдена в Excel файлі")
        
        sheet = self.workbook[sheet_name]
        data = []
        
        # Читаємо заголовки
        headers = []
        header_row_data = list(sheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True))
        if header_row_data:
            headers = [str(cell).strip() if cell else f"Column_{i+1}" 
                      for i, cell in enumerate(header_row_data[0])]
        
        # Читаємо дані
        for row_idx, row in enumerate(sheet.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
            # Пропускаємо порожні рядки
            if not any(cell for cell in row):
                continue
            
            row_data = {}
            for col_idx, cell_value in enumerate(row):
                header_name = headers[col_idx] if col_idx < len(headers) else f"Column_{col_idx+1}"
                row_data[header_name] = cell_value
            
            if row_data:
                data.append(row_data)
        
        return data
    
    def get_categories_data(self, sheet_name: str = "Категорія+") -> List[Dict[str, any]]:
        """
        Отримати дані категорій з вкладки "Категорія+"
        
        Args:
            sheet_name: Назва вкладки з категоріями (за замовчуванням "Категорія+")
        
        Returns:
            Список словників з даними категорій
            Очікувані колонки: ID категорії з фід, Назва категорії з фід, тощо
        """
        if not self.sheet_exists(sheet_name):
            raise ValueError(f"Вкладка '{sheet_name}' не знайдена в Excel файлі")
        
        return self.read_sheet_data(sheet_name)
    
    def get_category_id_and_name_from_feed(self, sheet_name: str = "Категорія+") -> List[Tuple[str, str]]:
        """
        Отримати список пар (ID категорії з фід, Назва категорії з фід) з вкладки.
        Підтримує різні варіанти назв колонок та заголовок у рядках 1 або 2.
        
        Returns:
            Список кортежів (category_id_from_feed, category_name_from_feed)
        """
        for header_row in (1, 2):
            categories_data = self.read_sheet_data(sheet_name, header_row=header_row)
            result = []
            id_key = None
            name_key = None
            
            if not categories_data:
                continue
            
            first_row = categories_data[0]
            keys_lower = {k.lower(): k for k in first_row.keys()}
            
            # Визначаємо колонку ID: містить "id" та (опційно) "фід"/"feed"/"категор"
            for k in keys_lower:
                if 'id' not in k:
                    continue
                if any(x in k for x in ('фід', 'feed', 'категор', 'category')):
                    id_key = keys_lower[k]
                    break
            if not id_key:
                for k in keys_lower:
                    if k in ('id', 'id категорії', 'category id', 'id з фід'):
                        id_key = keys_lower[k]
                        break
            if not id_key and len(keys_lower) >= 1:
                id_key = list(first_row.keys())[0]
            
            # Визначаємо колонку назви: "Назва категорії з фід" або "Категорії фіду" (друга колонка у файлі)
            for k in keys_lower:
                if ('назва' in k or 'name' in k) and any(x in k for x in ('фід', 'feed', 'категор', 'category')):
                    name_key = keys_lower[k]
                    break
            if not name_key:
                for k in keys_lower:
                    if k in ('назва', 'name', 'назва категорії', 'category name', 'категорії фіду', 'категорії фід'):
                        name_key = keys_lower[k]
                        break
            if not name_key:
                # "Категорії фіду" — друга колонка (назви категорій з фіду)
                for k in keys_lower:
                    if 'категор' in k and 'фід' in k and 'id' not in k:
                        name_key = keys_lower[k]
                        break
            if not name_key and len(keys_lower) >= 2:
                name_key = list(first_row.keys())[1]
            
            for row in categories_data:
                vid = row.get(id_key) if id_key else None
                vname = row.get(name_key) if name_key else None
                if vid is not None and vname is not None:
                    sid = str(vid).strip()
                    sname = str(vname).strip()
                    if sid and sname:
                        result.append((sid, sname))
            
            if result:
                return result
        
        # Останній fallback: читаємо сирі рядки (перші 2 колонки = id, назва)
        if not self.workbook or sheet_name not in self.workbook.sheetnames:
            return []
        sheet = self.workbook[sheet_name]
        result = []
        for row in sheet.iter_rows(min_row=1, values_only=True):
            if not row or len(row) < 2:
                continue
            c0, c1 = row[0], row[1]
            if c0 is None or c1 is None:
                continue
            s0, s1 = str(c0).strip(), str(c1).strip()
            if not s0 or not s1:
                continue
            # Пропускаємо рядок-заголовок
            s0_lower, s1_lower = s0.lower(), s1.lower()
            if s0_lower in ('id', 'назва', 'name', 'id категорії', 'назва категорії', 'id категорії фіду') or s1_lower in ('id', 'назва', 'name', 'категорії фіду'):
                continue
            if 'id' in s0_lower and 'категорії' in s0_lower and 'фід' in s0_lower:
                continue
            if 'категорії' in s1_lower and 'фід' in s1_lower and len(s1) < 25:
                continue
            result.append((s0, s1))
        return result
    
    def _load_xml_feed(self, xml_feed_url: str) -> ET.Element:
        """
        Завантажити XML фід з URL та повернути корінь XML дерева
        
        Args:
            xml_feed_url: URL XML фіду
        
        Returns:
            Корінь XML дерева (Element)
        
        Raises:
            Exception: Якщо не вдалося завантажити або розпарсити XML
        """
        try:
            response = requests.get(xml_feed_url, timeout=30)
            response.raise_for_status()
            xml_content = response.content
            root = ET.fromstring(xml_content)
            return root
        except Exception as e:
            raise Exception(f"Помилка при завантаженні XML фіду з URL '{xml_feed_url}': {e}")
    
    def _extract_categories_from_xml(self, xml_root: ET.Element) -> List[Tuple[str, str]]:
        """
        Витягти категорії з XML фіду
        
        Args:
            xml_root: Корінь XML дерева
        
        Returns:
            Список кортежів (category_id, category_name)
        """
        categories = []
        
        # Шукаємо всі елементи <category>
        # XML структура: <category id="1000" rz_id="169823">Домашній текстиль</category>
        # Категорії можуть бути в різних місцях XML (в <categories> або <shop><categories>)
        for category in xml_root.iter('category'):
            category_id = category.get('id')
            # Отримуємо текст категорії (може бути None або порожній рядок)
            category_text = category.text
            category_name = category_text.strip() if category_text and category_text.strip() else None
            
            # Додаємо тільки якщо є і ID і назва
            if category_id and category_name:
                categories.append((str(category_id).strip(), category_name))
        
        return categories
    
    def compare_categories_with_xml_feed(self, xml_feed_url: str, sheet_name: str = "Категорія+") -> Dict[str, any]:
        """
        Порівняти категорії з Excel файлу з категоріями з XML фіду
        
        Args:
            xml_feed_url: URL XML фіду для порівняння
            sheet_name: Назва вкладки з категоріями в Excel (за замовчуванням "Категорія+")
        
        Returns:
            Словник з результатами порівняння:
            - categories_match: чи відповідають категорії
            - missing_in_excel: категорії які є в XML але відсутні в Excel
            - missing_in_xml: категорії які є в Excel але відсутні в XML
            - excel_categories_count: кількість категорій в Excel
            - xml_categories_count: кількість категорій в XML
            - details: детальна інформація про порівняння
        """
        # Отримуємо категорії з Excel
        excel_categories = self.get_category_id_and_name_from_feed(sheet_name)
        excel_categories_dict = {cat_id: cat_name for cat_id, cat_name in excel_categories}
        
        # Завантажуємо та парсимо XML фід
        xml_root = self._load_xml_feed(xml_feed_url)
        xml_categories = self._extract_categories_from_xml(xml_root)
        xml_categories_dict = {cat_id: cat_name for cat_id, cat_name in xml_categories}
        
        # Порівнюємо категорії
        excel_ids = set(excel_categories_dict.keys())
        xml_ids = set(xml_categories_dict.keys())
        
        missing_in_excel = []
        missing_in_xml = []
        mismatched_names = []
        
        # Категорії які є в XML але відсутні в Excel
        for cat_id in xml_ids - excel_ids:
            missing_in_excel.append({
                "id": cat_id,
                "name": xml_categories_dict[cat_id]
            })
        
        # Категорії які є в Excel але відсутні в XML
        for cat_id in excel_ids - xml_ids:
            missing_in_xml.append({
                "id": cat_id,
                "name": excel_categories_dict[cat_id]
            })
        
        # Перевіряємо назви категорій для спільних ID
        # Excel може мати ієрархічний формат "Батько > Дочірня категорія", XML — лише назву
        def _names_match(excel_name: str, xml_name: str) -> bool:
            if excel_name == xml_name:
                return True
            # Excel: "Дім і сад > Домашній текстиль", XML: "Домашній текстиль"
            if excel_name.endswith(" > " + xml_name) or excel_name.strip().endswith(xml_name):
                return True
            return False
        
        common_ids = excel_ids & xml_ids
        for cat_id in common_ids:
            excel_name = excel_categories_dict[cat_id]
            xml_name = xml_categories_dict[cat_id]
            if not _names_match(excel_name, xml_name):
                mismatched_names.append({
                    "id": cat_id,
                    "excel_name": excel_name,
                    "xml_name": xml_name
                })
        
        categories_match = (
            len(missing_in_excel) == 0 and
            len(missing_in_xml) == 0 and
            len(mismatched_names) == 0
        )
        
        details = []
        if missing_in_excel:
            details.append(f"Відсутні в Excel ({len(missing_in_excel)}): {', '.join([c['id'] for c in missing_in_excel[:5]])}")
        if missing_in_xml:
            details.append(f"Відсутні в XML ({len(missing_in_xml)}): {', '.join([c['id'] for c in missing_in_xml[:5]])}")
        if mismatched_names:
            details.append(f"Невідповідність назв ({len(mismatched_names)}): {', '.join([c['id'] for c in mismatched_names[:5]])}")
        
        return {
            "categories_match": categories_match,
            "missing_in_excel": missing_in_excel,
            "missing_in_xml": missing_in_xml,
            "mismatched_names": mismatched_names,
            "excel_categories_count": len(excel_categories),
            "xml_categories_count": len(xml_categories),
            "common_categories_count": len(common_ids),
            "details": "; ".join(details) if details else "Всі категорії відповідають"
        }
    
    def close(self):
        """Закрити Excel файл"""
        if self.workbook:
            self.workbook.close()
            self.workbook = None
    
    def __enter__(self):
        """Контекстний менеджер - входження"""
        return self
    
    def __exit__(self, exc_type, exc_val, exc_tb):
        """Контекстний менеджер - вихід"""
        self.close()
